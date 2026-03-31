"""Excel logging for trade records."""

import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import EXCEL_LOG_PATH

logger = logging.getLogger(__name__)


class ExcelLogger:
    """Manages Excel trade logging with formatting."""

    def __init__(self, file_path: Path = EXCEL_LOG_PATH):
        """
        Initialize Excel logger.
        
        Args:
            file_path: Path to Excel file
        """
        self.file_path = file_path
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        self.init_workbook()

    def init_workbook(self):
        """Initialize Excel workbook with headers."""
        try:
            if self.file_path.exists():
                self.wb = openpyxl.load_workbook(str(self.file_path))
                self.ws = self.wb.active
            else:
                self.wb = openpyxl.Workbook()
                self.ws = self.wb.active
                self.ws.title = "Trades"
                self._create_headers()
                self.wb.save(str(self.file_path))
            
            logger.info(f"✓ Excel workbook initialized at {self.file_path}")
        except Exception as e:
            logger.error(f"✗ Failed to initialize Excel workbook: {e}")
            raise

    def _create_headers(self):
        """Create Excel headers with formatting."""
        headers = [
            'Timestamp', 'Strategy', 'Pair', 'Side', 'Entry Price', 'Exit Price',
            'Quantity', 'Leverage', 'Entry Time', 'Exit Time', 'Fees (INR)',
            'GST (18%)', 'Realized P&L (INR)', 'Tax (30%)', 'Cumulative P&L',
            'Status', 'Notes'
        ]

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        column_widths = [15, 15, 12, 8, 12, 12, 10, 8, 15, 15, 12, 10, 15, 10, 15, 10, 20]
        for col, width in enumerate(column_widths, 1):
            self.ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    def log_trade(self, trade_data: Dict) -> bool:
        """
        Log a trade to Excel.
        
        Args:
            trade_data: Dictionary with trade details
            
        Returns:
            True if successful, False otherwise
        """
        try:
            row = self.ws.max_row + 1
            
            values = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                trade_data.get('strategy_name', ''),
                trade_data.get('pair', ''),
                trade_data.get('side', ''),
                trade_data.get('entry_price', 0),
                trade_data.get('exit_price', 0),
                trade_data.get('quantity', 0),
                trade_data.get('leverage', 1.0),
                trade_data.get('entry_time', ''),
                trade_data.get('exit_time', ''),
                trade_data.get('fees_inr', 0),
                trade_data.get('gst', 0),
                trade_data.get('realized_pnl_inr', 0),
                trade_data.get('tax_estimate_30pct', 0),
                trade_data.get('cumulative_pnl', 0),
                trade_data.get('status', 'closed'),
                trade_data.get('notes', ''),
            ]

            for col, value in enumerate(values, 1):
                cell = self.ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Color code P&L
                if col == 13:  # Realized P&L column
                    if value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            self.wb.save(str(self.file_path))
            logger.info(f"✓ Trade logged to Excel: {trade_data.get('pair')}")
            return True
        except Exception as e:
            logger.error(f"✗ Failed to log trade to Excel: {e}")
            return False

    def add_summary_sheet(self, summary_data: Dict):
        """
        Add or update summary sheet with statistics.
        
        Args:
            summary_data: Dictionary with summary statistics
        """
        try:
            if "Summary" in self.wb.sheetnames:
                ws = self.wb["Summary"]
            else:
                ws = self.wb.create_sheet("Summary", 0)

            ws.clear()

            # Title
            ws['A1'] = "IndiaCryptoAlpha Trading Summary"
            ws['A1'].font = Font(bold=True, size=14)

            # Summary data
            row = 3
            for key, value in summary_data.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

            self.wb.save(str(self.file_path))
            logger.info("✓ Summary sheet updated")
        except Exception as e:
            logger.error(f"✗ Failed to add summary sheet: {e}")

    def get_trades(self, limit: int = 100) -> List[Dict]:
        """
        Read trades from Excel.
        
        Args:
            limit: Maximum number of trades to read
            
        Returns:
            List of trade dictionaries
        """
        try:
            trades = []
            headers = [cell.value for cell in self.ws[1]]
            
            for row in list(self.ws.iter_rows(min_row=2, values_only=True))[-limit:]:
                if row[0] is None:  # Skip empty rows
                    continue
                trade = {headers[i]: row[i] for i in range(len(headers))}
                trades.append(trade)
            
            return trades
        except Exception as e:
            logger.error(f"✗ Failed to read trades from Excel: {e}")
            return []

    def close(self):
        """Close workbook."""
        try:
            self.wb.close()
            logger.info("✓ Excel workbook closed")
        except Exception as e:
            logger.error(f"✗ Failed to close Excel workbook: {e}")
